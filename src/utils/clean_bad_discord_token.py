import os
import asyncio
import aiofiles
from openpyxl import load_workbook
from src.logger import AsyncLogger

file_lock = asyncio.Lock()
logger = AsyncLogger()

async def clean_bad_auth_tokens_discord(auth_tokens_discord: str) -> None:
    config_dir = os.path.join("config", "data", "client")
    bad_token_path = os.path.join(config_dir, "bad_token.txt")
    accounts_path = os.path.join(config_dir, "accounts.xlsx")
    
    if not os.path.exists(config_dir):
        os.makedirs(config_dir)
   
    async with file_lock:
        try:
            async with aiofiles.open(bad_token_path, 'a') as file:
                await file.write(f"{auth_tokens_discord}\n")
            
            if not os.path.exists(accounts_path):
                return

            try:
                wb = load_workbook(accounts_path, read_only=True)
                ws = wb.active
                
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                token_col_idx = next((i for i, cell in enumerate(header_row) if cell == 'Discord Token'), None)
                
                if token_col_idx is None:
                    wb.close()
                    return
                
                wb_write = load_workbook(accounts_path)
                ws_write = wb_write.active
                rows_modified = 0
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    if row[token_col_idx] == auth_tokens_discord:
                        ws_write.cell(row_idx, token_col_idx + 1, value="")
                        rows_modified += 1
                
                if rows_modified:
                    wb_write.save(accounts_path)
                    await logger.logger_msg(
                        f"Cleared bad Discord token from accounts.xlsx",
                        type_msg="info"
                    )
                
                wb.close()
                wb_write.close()
                
            except Exception as e:
                await logger.logger_msg(
                    f"Error updating accounts.xlsx: {str(e)}",
                    type_msg="error", method_name="clean_bad_auth_tokens_discord"
                )
                   
        except Exception as e:
            await logger.logger_msg(
                f"Error processing Discord token: {str(e)}",
                type_msg="error", method_name="clean_bad_auth_tokens_discord"
            )